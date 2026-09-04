#!/usr/bin/env python3
"""Mint a SpreadsheetBench task set — and REFUSE the cases no correct agent could have passed.

`sbench_digest.py` holds a salted digest of the answer so the oracle need not ride into the agent's
container. Whoever mints those digests decides what the exam IS, and a digest minted from a broken answer
key is an exam that scores a correct answer zero. Nothing downstream can tell that apart from an agent that
got it wrong: the reward file says 0.0 either way, the round is comparable, the campaign's gate derives a
verdict, and the whole loop optimizes inside a measurement that was never winnable.

So the mint and the check have ONE owner, and it is this file. It imports `parse_answer_position` and
`digest_of` from `sbench_digest` rather than restating them — a predicate written twice has already
diverged (rule `protocol`, L3).

WHAT IT REFUSES, and why each one is PROVABLE rather than suspected:

  · an `answer_position` this grader cannot read. 420 of the published 912 are sheet-qualified or
    multi-range; before `parse_answer_position` existed they raised inside the per-file read and were
    recorded as "output could not be read" — an ordinary 0.0. A question the grader cannot ask may not
    be answered (rule `protocol`, L2).
  · fewer than four answer cells. `sbench_digest.py`'s docstring already states this bound — a digest over
    one small number is guessable — and until now nothing enforced it.
  · a MISPAIRED answer key: comparing the cells OUTSIDE the answer range, no answer workbook pairs with its
    OWN input, and a bijection onto the other inputs does. Every answer matched to exactly one input and
    every input claimed once is a permutation; no task effect produces one. Case 15380 is one — answers 2
    and 3 are swapped — so an agent that solves all three workbooks correctly is scored 1/3 and fails.
    Measured, on the real data, with a real agent's real output.

    THE IDENTITY IS ASKED FIRST, and that ordering is the whole check. Two test cases commonly differ only
    INSIDE the answer range, which makes their contexts equal — and then the identity and a swap of those
    two are both bijections. Looking for a non-identity permutation and taking the first hit reported 70
    correct keys as mispaired, all naming the same swap, which is what gave it away.

WHAT IT CANNOT CHECK, and says so. Not every answer workbook is "the input with the answer written into
it": for an extraction task it is a result-only sheet with nothing outside `answer_position` at all. There
is then no evidence either way, and the honest answer is a third value rather than a verdict. The first
version of this file did not have it, compared `{}` to `{}`, and refused 10 cases on an empty match — case
97-36 has all 22 of its values inside `A1:A22`, so every answer file matched whichever input also happened
to have nothing outside. Those cases are staged and counted separately.

WHAT IT ONLY REPORTS. An answer workbook whose non-answer cells match NO input is not proof of anything: a
task that sorts a table or fills a column legitimately changes cells outside `answer_position`. Cases 17047
and 30930 are in this bucket and look like data slips on inspection (one literal in a column the
instruction never writes — `A31` is -3039 in the input and -30 in the answer), but "looks like" is not a
refusal. They are printed with the differing cells so a human can decide, and they are still minted.
"""
import argparse, hashlib, json, pathlib, sys, itertools
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from sbench_digest import canon, digest_of
from sbench_position import cell_count, parse_answer_position

MIN_ANSWER_CELLS = 4


def context(path, positions, fallback_sheet):
    """Every non-empty cell OUTSIDE the answer range, canonicalized the way the grader canonicalizes.

    The grader's own `canon` is reused so this comparison rounds floats exactly as scoring does — without
    that, `0.0009995002498750624` and `0.000999500249875062` read as a difference and every case with a
    computed column looks mispaired.
    """
    wb = load_workbook(path, data_only=True)
    covered = set()
    for part_sheet, rng in positions:
        name = part_sheet or fallback_sheet
        ws = wb[name] if (name and name in wb.sheetnames) else wb.active
        c1, r1, c2, r2 = range_boundaries(rng)
        covered |= {(ws.title, r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)}
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if (ws.title, cell.row, cell.column) in covered or cell.value is None:
                    continue
                key = canon(cell)
                if key != "S":  # an empty string is an absent value, not a difference
                    out[(ws.title, cell.coordinate)] = key
    return out


def examine(root, task, salt):
    """-> (kind, detail, digests).

    kind is 'admitted' (the key pairs with its input), 'refused' (provably unwinnable), 'no_evidence' (the
    answer workbooks carry nothing the pairing could be checked against) or 'admitted_with_report' (they
    differ, which a task legitimately can cause). The last two are staged; only 'refused' is dropped.
    """
    cid = str(task["id"])
    directory = root / cid
    try:
        positions = parse_answer_position(task["answer_position"])
    except Exception as exc:
        return "refused", f"answer_position {task['answer_position']!r} is unreadable ({exc})", None
    cells = cell_count(positions)
    if cells < MIN_ANSWER_CELLS:
        return "refused", f"{cells} answer cell(s): a digest over so few values is guessable", None

    sheet = task.get("answer_sheet") or ""
    try:
        inputs = {n: context(directory / f"{n}_{cid}_input.xlsx", positions, sheet) for n in (1, 2, 3)}
        answers = {n: context(directory / f"{n}_{cid}_answer.xlsx", positions, sheet) for n in (1, 2, 3)}
    except Exception as exc:
        return "refused", f"a workbook could not be read ({type(exc).__name__}: {exc})", None

    def digests_for():
        return [digest_of(directory / f"{n}_{cid}_answer.xlsx", sheet, positions, salt) for n in (1, 2, 3)]

    # AN EMPTY CONTEXT MATCHES AN EMPTY CONTEXT, AND THAT IS NOT EVIDENCE OF ANYTHING. Not every answer
    # workbook is "the input with the answer filled in": for an extraction task it is a result-only sheet,
    # so everything outside `answer_position` is absent from it. Case 97-36 is that shape — all 22 of its
    # values live inside `A1:A22`, every answer file has nothing outside it, and `{} == {}` reported the
    # key as mispaired against whichever input also happened to have nothing outside. The first version of
    # this check refused 11 cases and 10 of them were this. So a pairing claim needs discriminating bytes.
    if not all(answers[n] for n in (1, 2, 3)):
        return "no_evidence", "the answer workbooks hold nothing outside the answer range", digests_for()

    matches = {n: [m for m in (1, 2, 3) if answers[n] == inputs[m]] for n in (1, 2, 3)}

    # THE IDENTITY IS ASKED FIRST, AND IT ANSWERS THE QUESTION. Two test cases often differ only inside the
    # answer range, which makes their contexts equal — and then the identity AND a swap of those two are
    # both bijections. Searching for a non-identity permutation and taking the first hit reported 70 cases
    # as "answer 2 is input 3's, answer 3 is input 2's" whose keys were perfectly correct; the giveaway was
    # that all 70 named the same permutation. A key that pairs with its own input is paired, however many
    # other readings the data also admits.
    if all(n in matches[n] for n in (1, 2, 3)):
        return "admitted", "", digests_for()

    # Only now is a bijection evidence: every answer matched to exactly one input, every input claimed
    # once, and no self-pairing available. That is a permutation, and no task effect produces one.
    permutation = next(
        (p for p in itertools.permutations((1, 2, 3))
         if p != (1, 2, 3) and all(answers[n] == inputs[p[n - 1]] for n in (1, 2, 3))),
        None,
    )
    if permutation is not None:
        pairs = ", ".join(f"answer {n} is input {permutation[n - 1]}'s" for n in (1, 2, 3)
                          if permutation[n - 1] != n)
        return "refused", f"the answer key is mispaired ({pairs})", None

    digests = digests_for()
    drifted = [n for n in (1, 2, 3) if n not in matches[n]]
    if drifted:
        detail = []
        for n in drifted:
            diff = [k for k in set(answers[n]) | set(inputs[n]) if answers[n].get(k) != inputs[n].get(k)]
            shown = ", ".join(f"{s}!{c}: input={inputs[n].get((s, c))} answer={answers[n].get((s, c))}"
                              for s, c in sorted(diff)[:4])
            detail.append(f"file {n} differs outside the answer range at {len(diff)} cell(s) — {shown}")
        return "admitted_with_report", "; ".join(detail), digests
    return "admitted", "", digests


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True, help="the unpacked all_data_912 directory")
    ap.add_argument("--out", required=True, help="where to write tasks.json")
    ap.add_argument("--ids", default="", help="comma-separated instruction ids; default is every one")
    args = ap.parse_args()

    root = pathlib.Path(args.data)
    dataset = json.loads((root / "dataset.json").read_text())
    if args.ids:
        wanted = {i.strip() for i in args.ids.split(",") if i.strip()}
        dataset = [t for t in dataset if str(t["id"]) in wanted]
        missing = wanted - {str(t["id"]) for t in dataset}
        if missing:
            print(f"no such instruction: {', '.join(sorted(missing))}", file=sys.stderr)
            return 2

    staged, refused, reported, unchecked = [], [], [], []
    for task in dataset:
        cid = str(task["id"])
        salt = hashlib.sha256(f"{cid}|everdict-spreadsheetbench".encode()).hexdigest()[:16]
        kind, detail, digests = examine(root / "spreadsheet", task, salt)
        if kind == "refused":
            refused.append((cid, detail))
            continue
        if kind == "admitted_with_report":
            reported.append((cid, detail))
        if kind == "no_evidence":
            unchecked.append(cid)
        staged.append({
            "id": task["id"],
            "instruction": task["instruction"],
            "answer_position": task["answer_position"],
            "answer_sheet": task.get("answer_sheet", ""),
            "salt": salt,
            "digests": digests,
        })

    pathlib.Path(args.out).write_text(json.dumps(staged, indent=2))
    print(f"staged {len(staged)} of {len(dataset)} instructions -> {args.out}")
    if unchecked:
        print(f"{len(unchecked)} of them could not have their answer key checked at all — the answer "
              f"workbooks hold nothing outside the answer range, so there is no pairing evidence either way")
    if reported:
        print(f"\n{len(reported)} admitted with a pairing report (a human should read these):", file=sys.stderr)
        for cid, detail in reported:
            print(f"  {cid}: {detail}", file=sys.stderr)
    if refused:
        print(f"\n{len(refused)} REFUSED — unwinnable as published, so they are not an exam:", file=sys.stderr)
        for cid, detail in refused:
            print(f"  {cid}: {detail}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
