#!/usr/bin/env python3
"""Self-inspection for a produced SpreadsheetBench workbook — what the READER will see, and nothing else.

THE ENVIRONMENT'S TOOL, NOT THE AGENT'S CLEVERNESS. Measuring a small model on this benchmark showed two
failure modes that dominate and that have nothing to do with spreadsheet reasoning:

  · an answer column filled for the first N rows and left empty for the rest (24 of 49, identically in all
    three workbooks — the agent believed it was finished);
  · a formula the reader cannot evaluate, so every cell reads `None` after recalculation, which is exactly
    how the official eval reads it (`data_only=True`).

Both are decidable from the agent's OWN output and its OWN input. This tool makes that decision cheap and
reliable instead of leaving each agent to re-derive it, which is what a harness is for.

IT CANNOT LEAK AN ANSWER: it opens only the files already in the container the agent works in, and reports
shape — where the data ends, which cells are empty, which formulas failed to evaluate. It never sees the
answer workbook (which is not in the image at all) and never reports whether a value is correct.
"""
import argparse, sys
from openpyxl import load_workbook


def report(path: str) -> int:
    try:
        cached = load_workbook(path, data_only=True)
        live = load_workbook(path, data_only=False)
    except Exception as exc:
        print(f"{path}: cannot be opened ({type(exc).__name__}: {exc})", file=sys.stderr)
        return 1

    problems = 0
    for name in cached.sheetnames:
        cs, ls = cached[name], live[name]
        extent = max((c.row for row in cs.iter_rows() for c in row if c.value is not None), default=0)
        print(f"[{name}] rows with data: {extent}, columns: {cs.max_column}")
        for col in range(1, cs.max_column + 1):
            filled, empty_after, dead = [], [], []
            for row in range(1, extent + 1):
                cached_value = cs.cell(row=row, column=col).value
                live_value = ls.cell(row=row, column=col).value
                is_formula = isinstance(live_value, str) and live_value.startswith("=")
                if cached_value is not None:
                    filled.append(row)
                elif is_formula:
                    dead.append(row)  # a formula the reader could not evaluate → reads as empty
                elif filled:
                    empty_after.append(row)
            if not filled and not dead:
                continue
            letter = cs.cell(row=1, column=col).column_letter
            head = str(cs.cell(row=1, column=col).value or "")[:20]
            line = f"  {letter} ({head or 'no header'}): {len(filled)} value(s) up to row {max(filled, default=0)}"
            if empty_after:
                line += f"  ⚠ EMPTY at rows {empty_after[:6]}{'…' if len(empty_after) > 6 else ''} while data runs to {extent}"
                problems += 1
            if dead:
                line += f"  ⚠ {len(dead)} formula(s) produced NO value (rows {dead[:6]}) — the reader sees these as empty"
                problems += 1
            print(line)
    if problems:
        print(f"\n{problems} shape problem(s): the reader would see empty cells where this sheet has data.", file=sys.stderr)
    else:
        print("\nno shape problems: every column with data is filled to the sheet's extent and every formula evaluated.")
    return 2 if problems else 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("files", nargs="+")
    args = ap.parse_args()
    worst = 0
    for path in args.files:
        print(f"=== {path} ===")
        worst = max(worst, report(path))
    return worst


if __name__ == "__main__":
    sys.exit(main())
