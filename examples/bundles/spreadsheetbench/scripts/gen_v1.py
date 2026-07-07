# Generate a SpreadsheetBench v1-style sample input — plants a real xlsx spreadsheet in the working directory.
# Real SpreadsheetBench (912 instructions) ships per-case xlsx collected from forums, but a bundle cannot inline binaries,
# so the sample generates a deterministic xlsx via this script in setup (for demonstrating the pipeline).
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales"
ws["A1"], ws["B1"] = "Region", "Amount"
rows = [("North", 120), ("South", 340), ("East", 80), ("West", 260), ("North", 150)]
for i, (region, amount) in enumerate(rows, start=2):
    ws[f"A{i}"], ws[f"B{i}"] = region, amount
wb.save("input.xlsx")
print(f"wrote input.xlsx ({len(rows)} rows), sum(Amount)={sum(a for _, a in rows)}")
