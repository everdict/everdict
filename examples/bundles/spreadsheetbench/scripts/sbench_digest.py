#!/usr/bin/env python3
"""Score a SpreadsheetBench V1 task against a SALTED DIGEST of its answer instead of the answer workbook.

The official comparison is unchanged — cached cell values at `answer_position`, 2-dp rounding, exact type, and
all three test cases must pass. What changes is what the checker HOLDS: a sha256 of the canonicalized answer
under a per-task salt, not the answer itself.

WHY: an agent and its grader share one container on every runtime that has no separate verifier lane (only the
Nomad and K8s backends implement `dispatchVerifier`). Shipping the answer workbooks into that container would
put the oracle inside the thing being measured. A digest cannot be read back into values for a range of more
than a couple of cells, so the environment can carry the check without carrying the answer.

Its limit, stated: a digest over ONE small number is guessable, so tasks are selected with at least four
answer cells. This is a substitute for a private verifier, not a replacement for one.

`answer_position` IS NOT A PLAIN RANGE, AND READING IT AS ONE SCORED 417 OF 912 TASKS ZERO. This file used
to hand the whole published string to `range_boundaries`, which parses `C2:C66` and raises on every
sheet-qualified or multi-range form — and the raise landed in the per-file `except Exception` below, was
recorded as "test n: output could not be read", and became an ordinary 0.0 reward. 492 of the 912 published
positions are plain ranges; the other 420 were therefore unwinnable by construction and reported as an agent
that got the answer wrong, with the agent's workbook never opened. Observed before the fix, against the
OFFICIAL answer workbooks — which pass under the plain range:

    $ sbench_digest.py --id 30930 --range "'Sheet1'!C2:C66" …
    test 1: 1_30930_output.xlsx could not be read (ValueError)
    FAIL 3/3 test case(s)

`sbench_position.parse_answer_position` now reads that field for every scorer here (909 of 912 readable; the
remaining 3 are unscoreable as published, and it says so). The exit code says which question was answered:
0 the agent matched, 1 it did not, 2 the grader could not ask — L2 of rule `protocol`, that "we could not
find out" may never be spent as somebody else's verdict.
"""
import argparse, hashlib, sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from sbench_position import parse_answer_position

GRADER_CANNOT_RUN = 2


def canon(cell):
    v = cell.value
    if v is None:
        return "N"
    if isinstance(v, bool):
        return f"B{v}"
    if isinstance(v, (int, float)):
        return f"F{round(float(v), 2):.2f}"
    return "S" + str(v).strip()


def digest_of(path, sheet, positions, salt):
    """`positions` is `parse_answer_position`'s answer — never a raw `answer_position` string."""
    wb = load_workbook(path, data_only=True)
    parts = []
    for part_sheet, rng in positions:
        name = part_sheet or sheet
        ws = wb[name] if (name and name in wb.sheetnames) else wb.active
        lo_c, lo_r, hi_c, hi_r = range_boundaries(rng)
        parts += [canon(ws.cell(row=r, column=c)) for r in range(lo_r, hi_r + 1) for c in range(lo_c, hi_c + 1)]
    return hashlib.sha256((salt + "|" + "\x1f".join(parts)).encode()).hexdigest()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--id", required=True)
    ap.add_argument("--range", required=True)
    ap.add_argument("--sheet", default="")
    ap.add_argument("--salt", required=True)
    ap.add_argument("--digests", required=True, help="comma-separated, one per test case")
    args = ap.parse_args()

    try:
        positions = parse_answer_position(args.range)
    except Exception as exc:
        # NOT a failed test case. The grader cannot ask the question, so it may not answer it — see the
        # module docstring, and L2 of rule `protocol`.
        print(f"grader cannot run: answer_position {args.range!r} is unreadable ({exc})", file=sys.stderr)
        return GRADER_CANNOT_RUN

    want = args.digests.split(",")
    failed = []
    for n, expected in enumerate(want, start=1):
        out = f"{n}_{args.id}_output.xlsx"
        try:
            got = digest_of(out, args.sheet, positions, args.salt)
        except FileNotFoundError:
            failed.append(f"test {n}: {out} was not produced")
            continue
        except Exception as exc:  # a workbook that cannot be read is a failed test case, not a crashed grader
            failed.append(f"test {n}: {out} could not be read ({type(exc).__name__})")
            continue
        if got != expected:
            failed.append(f"test {n}: {args.range} does not match")
    if failed:
        print("\n".join(failed), file=sys.stderr)
        print(f"FAIL {len(failed)}/{len(want)} test case(s)", file=sys.stderr)
        return 1
    print(f"PASS all {len(want)} test cases")
    return 0


if __name__ == "__main__":
    sys.exit(main())
