#!/usr/bin/env python3
# Port of SpreadsheetBench's official scoring logic (openpyxl only) — for scoring real data (golden xlsx).
# Run: python3 sbench_grade.py --version v1|v2 --output OUT.xlsx --golden GOLD.xlsx --answer-position "'Sheet'!A1:B10,C3" [--input IN.xlsx]
#   exit 0 = PASS, 1 = FAIL, 2 = error (missing file/openpyxl).
#   v1: compare answer_position cells value-wise (round to 2dp, type match), all must match for PASS.
#   v2: diff input→golden and score split into regression (invariant cells preserved) · modification (changed cells correct) (1% relative tolerance);
#       both regression and modification must be 100% for PASS (--input required, to split regression/modification).
# Note: the official eval reads cached values with data_only, so the produced xlsx must be pre-recalculated (LibreOffice, etc.).
import argparse
import datetime
import re
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed — cannot score", file=sys.stderr)
    sys.exit(2)

_RANGE_RE = re.compile(r"^(?:'([^']+)'!|([^'!]+)!)?(.+)$")


def parse_answer_position(spec):
    # Split on commas but ignore commas inside quotes. Each piece → (sheet|None, a1range).
    parts, buf, in_q = [], [], False
    for ch in spec:
        if ch == "'":
            in_q = not in_q
            buf.append(ch)
        elif ch == "," and not in_q:
            parts.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append("".join(buf).strip())
    out = []
    for p in parts:
        if not p:
            continue
        m = _RANGE_RE.match(p)
        sheet = m.group(1) or m.group(2)
        out.append((sheet.strip() if sheet else None, m.group(3).strip()))
    return out


def find_sheet(wb, name):
    if name is None:
        return wb[wb.sheetnames[0]]
    norm = lambda s: "".join(s.split()).lower()
    for sn in wb.sheetnames:
        if norm(sn) == norm(name):
            return wb[sn]
    raise KeyError(name)


def cells_of(ws, a1):
    if ":" not in a1:
        return [ws[a1]]
    return [c for row in ws[a1] for c in row]


def transform_value(v):
    # v1 normalization: numbers to 2dp, time/date, and float-parsable strings.
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.datetime):
        return round((v - datetime.datetime(1899, 12, 30)).days + 0.0, 0)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def eq_v1(a, b):
    ta, tb = transform_value(a), transform_value(b)
    if ta is None and tb is None:
        return True
    if type(ta) is not type(tb):
        return False
    return ta == tb


_NM = {"#div/0!", "#n/a", "n/a", "na", "n.a.", "n/m", "nm", "n.m.", "-", "--", "---", "–", "—"}


def eq_v2(a, b, tol=0.01):
    # Numbers within 1% relative tolerance, meaningless placeholders equivalent, None≡0, ""≡None.
    if (a is None or a == "") and (b is None or b == ""):
        return True
    if a is None:
        a = 0
    if b is None:
        b = 0
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        denom = max(abs(a), abs(b))
        return abs(a - b) <= (tol * denom if denom else tol)
    sa, sb = str(a).strip().lower(), str(b).strip().lower()
    if sa in _NM and sb in _NM:
        return True
    fa = sa.replace("$", "").lstrip("=+").replace("=", "", 1) if sa.startswith(("=", "+")) else sa
    fb = sb.replace("$", "").lstrip("=+").replace("=", "", 1) if sb.startswith(("=", "+")) else sb
    return fa == fb


def collect(path, positions, data_only=True):
    wb = openpyxl.load_workbook(path, data_only=data_only)
    vals = {}
    for sheet, a1 in positions:
        ws = find_sheet(wb, sheet)
        # The cell key uses answer_position's sheet designation (name | first sheet), not the "resolved sheet title" — so that even if
        # golden/output have different first-sheet titles (e.g. golden='Sheet', output='Sales'), the first sheets are compared positionally. The official eval is also first-sheet based.
        skey = sheet if sheet is not None else "\x00first"
        for c in cells_of(ws, a1):
            vals[(skey, c.coordinate)] = c.value
    return vals


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--version", choices=["v1", "v2"], required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--golden", required=True)
    ap.add_argument("--answer-position", required=True)
    ap.add_argument("--input", help="the original input xlsx needed to split v2 regression/modification")
    a = ap.parse_args()
    pos = parse_answer_position(a.answer_position)
    try:
        gold = collect(a.golden, pos)
        out = collect(a.output, pos)
    except FileNotFoundError as e:
        print(f"file not found: {e}")
        sys.exit(1)

    if a.version == "v1":
        bad = [k for k in gold if not eq_v1(gold[k], out.get(k))]
        print(f"v1: {len(gold)} cells, {len(bad)} mismatched -> {'PASS' if not bad else 'FAIL'}")
        if bad[:5]:
            print("  e.g.", [(k, gold[k], out.get(k)) for k in bad[:5]])
        sys.exit(0 if not bad else 1)

    # v2: split regression (invariant) vs modification (changed)
    if not a.input:
        print("v2 scoring requires --input (to split regression/modification)")
        sys.exit(2)
    inp = collect(a.input, pos)
    reg = [k for k in gold if eq_v2(inp.get(k), gold[k])]
    mod = [k for k in gold if k not in reg]
    reg_ok = sum(1 for k in reg if eq_v2(gold[k], out.get(k)))
    mod_ok = sum(1 for k in mod if eq_v2(gold[k], out.get(k)))
    reg_ratio = 1.0 if not reg else reg_ok / len(reg)
    mod_ratio = 1.0 if not mod else mod_ok / len(mod)
    if reg_ratio >= 0.998:
        reg_ratio = 1.0
    ok = reg_ratio == 1.0 and mod_ratio == 1.0
    print(f"v2: regression={reg_ok}/{len(reg)} modification={mod_ok}/{len(mod)} -> {'PASS' if ok else 'FAIL'}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
