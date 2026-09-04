#!/usr/bin/env python3
"""Check a produced workbook against the worked examples ALREADY IN ITS OWN INPUT — no answer required.

These tasks come from questions people asked about their own spreadsheet, and the asker usually filled in the
expected result for some rows by hand before posting. Those rows are in the INPUT the agent is given, so a
rule that disagrees with them is provably wrong — and provably wrong from what the agent already holds.

WHY A TOOL RATHER THAN AN INSTRUCTION. A scaffold that said "check your formula against the worked examples"
was measured on two tasks whose examples cover 100% of the answer range, and both still failed: telling an
agent to verify is not the same as giving it something that verifies. This reads both files and answers.

IT CANNOT LEAK AN ANSWER: it opens the agent's own input and the agent's own output, both already in the
container, and reports only where the two disagree on rows the input already had. It never sees an answer
workbook — there is none in the image — and it says nothing about the rows the asker left empty, which are
exactly the ones being graded.
"""
import argparse, sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def canon(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return f"B{v}"
    if isinstance(v, (int, float)):
        return f"F{round(float(v), 2):.2f}"
    return "S" + str(v).strip()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--range", required=True)
    ap.add_argument("--sheet", default="")
    args = ap.parse_args()

    try:
        src = load_workbook(args.input, data_only=True)
        got = load_workbook(args.output, data_only=True)
    except Exception as exc:
        print(f"cannot open: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1
    pick = lambda wb: wb[args.sheet] if (args.sheet and args.sheet in wb.sheetnames) else wb.active
    a, b, c, d = range_boundaries(args.range)
    si, so = pick(src), pick(got)

    examples, agree, disagree = 0, 0, []
    for y in range(b, d + 1):
        for x in range(a, c + 1):
            want = canon(si.cell(row=y, column=x).value)
            if want is None:
                continue  # the asker left this one for you — it is what you are being graded on
            examples += 1
            mine = canon(so.cell(row=y, column=x).value)
            if mine == want:
                agree += 1
            else:
                disagree.append((so.cell(row=y, column=x).coordinate, mine, want))

    if examples == 0:
        print("this sheet carries no worked examples — there is nothing here to check your rule against.")
        return 0
    print(f"worked examples in {args.input}: {examples}; your output reproduces {agree}.")
    if disagree:
        for coord, mine, want in disagree[:10]:
            print(f"  {coord}: you produced {mine}, the sheet already says {want}", file=sys.stderr)
        print(
            f"\n{len(disagree)} of {examples} worked example(s) disagree — your rule is wrong, "
            "and it is wrong in a way this sheet can already prove.",
            file=sys.stderr,
        )
        return 2
    print("every worked example agrees. That is necessary, not sufficient: the graded rows are the empty ones.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
