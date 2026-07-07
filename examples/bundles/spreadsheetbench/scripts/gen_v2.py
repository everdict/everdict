# Generate a SpreadsheetBench v2-sample (workflow/financial-model style) input — model.xlsx with a Revenue/Cost table.
# Real v2 Financial_Model is a multi-sheet model (several MB), but the sample reproduces v2's core scoring — regression (invariant cells preserved)
# + modification (new cells correct) — with a small table.
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model"
ws["A1"], ws["B1"], ws["C1"] = "Product", "Revenue", "Cost"
rows = [("Alpha", 1000, 600), ("Beta", 750, 500), ("Gamma", 1200, 900), ("Delta", 300, 100)]
for i, (p, r, c) in enumerate(rows, start=2):
    ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = p, r, c
wb.save("model.xlsx")
print(f"wrote model.xlsx ({len(rows)} rows); profits={[r - c for _, r, c in rows]} total={sum(r - c for _, r, c in rows)}")
