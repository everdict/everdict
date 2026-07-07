# SpreadsheetBench v2-sample scoring (recalculation) — verifies both regression + modification, which are v2's core.
#   modification: new column D 'Profit' = Revenue-Cost (per row) + a total-profit cell somewhere is correct.
#   regression:   are the original A/B/C columns (Product/Revenue/Cost) preserved as-is (invariant cells not corrupted)?
# Both must be 100% for PASS (v2 all-or-nothing). exit 0=PASS, 1=FAIL, 2=error.
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed — the pip install in setup failed", file=sys.stderr)
    sys.exit(2)

if not os.path.exists("output.xlsx"):
    print("output.xlsx is missing — the agent did not produce an output.")
    sys.exit(1)

src = openpyxl.load_workbook("model.xlsx")["Model"]
data = [
    (c[0].value, c[1].value, c[2].value)
    for c in zip(src["A"][1:], src["B"][1:], src["C"][1:])
    if isinstance(c[1].value, (int, float))
]
exp_profits = [r - co for _, r, co in data]
exp_total = sum(exp_profits)

ws = openpyxl.load_workbook("output.xlsx", data_only=True).active

# modification: is the per-row Profit (D2..) correct?
got_profits = [ws[f"D{i}"].value for i in range(2, 2 + len(exp_profits))]
mod_profit_ok = all(isinstance(g, (int, float)) and abs(g - e) < 1e-6 for g, e in zip(got_profits, exp_profits))
# modification: does a total-profit cell exist somewhere?
all_nums = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))]
mod_total_ok = any(abs(v - exp_total) < 1e-6 for v in all_nums)
# regression: are the original Revenue/Cost preserved (columns B/C)?
reg_ok = all(
    ws[f"B{i}"].value == r and ws[f"C{i}"].value == co for i, (_, r, co) in enumerate(data, start=2)
)

ok = mod_profit_ok and mod_total_ok and reg_ok
print(
    f"modification: profits={got_profits}(ok={mod_profit_ok}) total(ok={mod_total_ok}); "
    f"regression: Revenue/Cost preserved(ok={reg_ok}) -> {'PASS' if ok else 'FAIL'}"
)
sys.exit(0 if ok else 1)
